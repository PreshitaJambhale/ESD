
import openpyxl
from openpyxl import Workbook, load_workbook



class utils():

    @staticmethod
    def read_data_from_excel(file, sheet):
        data_list = []
        wb = load_workbook(filename=file)
        sh = wb[sheet]

        for i in range(2, sh.max_row + 1):
            row = []
            for j in range(1, sh.max_column + 1):
                row.append(sh.cell(row=i, column=j).value)

            if all(cell is None for cell in row):
                continue

            data_list.append(tuple(row))
            print(data_list)

        return data_list

    read_data_from_excel("/Users/harshalijambhale/Downloads/test sheets/subscription_stf_create_data.xlsx", "Sheet1")












    """

    @staticmethod
    def read_data_from_excel(file, sheet):

        data_list = []
        wb = load_workbook(filename = file)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct +1):
            row = []
            for j in range(1, col_ct +1):
                row.append(sh.cell(row =i, column =j ).value)

            if all(cell is None for cell in row):
                continue

            data_list.append(tuple(row))

        print(data_list)

        for item in data_list:
            if isinstance(item, tuple):
                data_list.remove(item)
                data_list.extend(item)
                break

        print(data_list)
        return data_list


    read_data_from_excel("/Users/harshalijambhale/eclipse-workspace/test/src/test/resources/testngData.xlsx", "Sheet1")

    data = [1, 3, 4, (2, 5)]

    for item in data:
        if isinstance(item, tuple):
            data.remove(item)
            data.extend(item)
            break

    print(data)
"""




